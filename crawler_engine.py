import os
import re
import time
import json
import threading
import base64
import csv
from datetime import datetime, timedelta
from urllib.parse import quote_plus
from urllib.parse import urlparse
from urllib.parse import parse_qs
from html import unescape

import openpyxl
import requests

class CrawlerEngine:
    def __init__(self):
        self.is_running = False
        self._stop_event = threading.Event()
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.run_config = self._load_run_config()
        self.config = {
            "location_name": "成武",
            "location_id": "152",
            "nodes": ["节点1"],
            "industry_chains": [],
            "news_days": 365,
            "search_results_per_query": 5,
            "max_urls_per_chain": 20,
            "max_nodes_per_chain": 6,
            "news_source": "raw_news_csv",
            "raw_news_days": 30,
            "max_raw_news_per_node": 30,
        }
        self.logs = []
        self.leads = []
        self.raw_news = []
        self.stats = {"total_crawled": 0, "valid_leads": 0, "raw_news": 0}
        self.nodes_file = ""
        self.node_meta = {}
        
        # 定义搜索组别 (依据招商情报专家规范)
        self.keyword_groups = {
            "组A（投资与扩张）": ["对外投资", "扩产", "搬迁", "新建工厂", "设立分公司", "产能扩张", "项目落户", "企业落户"],
            "组B（资本与融资）": ["融资", "正在融资", "拟融资", "股权融资", "IPO", "并购重组", "项目融资", "获得融资"],
            "组C（招商考察与对接）": ["代表团走访对接", "招商小分队", "上门拜访", "考察调研", "意向投资签约", "框架协议签约"],
            "组D（经营与资产）": ["产能短缺", "订单激增", "土地摘牌", "厂房建设", "固定资产投资", "设备采购"],
            "组E（政策与荣誉）": ["政府扶持项目", "高新技术企业扩产", "专精特新企业扩张", "获得技改补贴"],
            "组F（媒体与宣发特征词）": ["重大项目集中开工", "集中签约", "招大引强", "产业链链长制"]
        }

        self.deepseek_api_key = str(self._cfg_raw("DEEPSEEK_API_KEY", "deepseek_api_key", "") or "").strip()
        self.deepseek_api_base_url = str(self._cfg_raw("DEEPSEEK_API_BASE_URL", "deepseek_api_base_url", "https://api.deepseek.com") or "").strip()
        self.deepseek_model = str(self._cfg_raw("DEEPSEEK_MODEL", "deepseek_model", "deepseek-chat") or "").strip()
        self.deepseek_timeout_seconds = int(self._cfg_raw("DEEPSEEK_TIMEOUT_SECONDS", "deepseek_timeout_seconds", 120) or 120)

    def _normalize_text(self, s):
        if s is None:
            return ""
        return re.sub(r"\s+", " ", str(s).strip())

    def _parse_date_ymd(self, s):
        t = self._normalize_text(s)
        if not t:
            return None
        m = re.search(r"(\d{4})[-/\.](\d{1,2})[-/\.](\d{1,2})", t)
        if not m:
            return None
        try:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return datetime(y, mo, d).date()
        except Exception:
            return None

    def _find_raw_news_csv(self):
        candidates = []
        wd = os.getcwd()
        candidates.append(os.path.join(wd, "raw_news_leads.csv"))
        candidates.append(os.path.join(wd, "raw_news.csv"))
        parent = os.path.dirname(self.base_dir)
        candidates.append(os.path.join(parent, "raw_news_leads.csv"))
        candidates.append(os.path.join(parent, "raw_news.csv"))
        for p in candidates:
            if p and os.path.exists(p):
                return p
        return ""

    def _load_raw_news_rows_from_csv(self, path):
        p = (path or "").strip()
        if not p or not os.path.exists(p):
            return []
        rows = []
        header_map = {
            "鎵€灞炰骇涓氶摼": "所属产业链",
            "鎵€灞炰骇涓?": "所属产业",
            "鎵€灞炰骇涓�": "所属产业",
            "鎼滅储缁勫埆": "搜索组别",
            "鏂伴椈鏍囬": "新闻标题",
            "鏂伴椈鎽樿": "新闻摘要",
            "鍏抽敭浜嬩欢": "关键事件",
            "娑夊強浼佷笟/鏈烘瀯": "涉及企业/机构",
            "鏂伴椈閾炬帴": "新闻链接",
            "淇℃伅鏉ユ簮骞冲彴": "信息来源平台",
            "鍙戝竷鏃堕棿": "发布时间",
            "鍙戝竷 鏃堕棿": "发布时间",
            "鎶撳彇鏃堕棿": "抓取时间",
        }

        try:
            with open(p, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for r in reader:
                    if not isinstance(r, dict):
                        continue
                    rr = {}
                    for k, v in r.items():
                        kk = self._normalize_text(k).replace(" ", "")
                        kk = header_map.get(kk, kk)
                        rr[kk] = v
                    r = rr
                    title = self._normalize_text(r.get("新闻标题") or "")
                    summary = self._normalize_text(r.get("新闻摘要") or "")
                    link = self._normalize_text(r.get("新闻链接") or "")
                    if not title or not summary or not link:
                        continue
                    rows.append(r)
        except Exception:
            return []

        days = int(self.config.get("raw_news_days") or 30)
        cutoff = (datetime.now().date() - timedelta(days=days))
        filtered = []
        for r in rows:
            pub = self._parse_date_ymd(r.get("发布时间") or "")
            if pub and pub < cutoff:
                continue
            filtered.append(r)
        return filtered

    def _build_penetration_prompt_from_raw_news(self, chain_name, node_name, group_name, keyword, raw_row, meta):
        loc_name = self.config.get("location_name") or "未披露"
        chain_ids = str((meta or {}).get("chain_ids") or "")
        node_ids = str((meta or {}).get("node_ids") or "")
        chain_node = f"{chain_name}-{node_name}"

        title = self._normalize_text((raw_row or {}).get("新闻标题") or "")
        summary = self._normalize_text((raw_row or {}).get("新闻摘要") or "")
        entities = self._normalize_text((raw_row or {}).get("涉及企业/机构") or "")
        key_event = self._normalize_text((raw_row or {}).get("关键事件") or "")
        link = self._normalize_text((raw_row or {}).get("新闻链接") or "")
        publish_date = self._normalize_text((raw_row or {}).get("发布时间") or "") or "N/A"
        platform = self._normalize_text((raw_row or {}).get("信息来源平台") or "")

        return f"""你现在是一名“招商线索穿透分析师”。我会提供一条【原始新闻舆情】（来自 crawl_raw_news.py 的广度采集结果），请基于标题+摘要+事件等信息，穿透提取可落地的招商线索。\n\n硬性规则：\n1) company_name 必须是真实工商企业全称，且必须以“有限公司/股份有限公司/有限责任公司/集团有限公司”结尾；严禁使用“某公司/某企业/签约企业/未披露”等代称；无法确认输出 []\n2) public_sentiment_link 必须使用我提供的新闻链接 link，且不得伪造\n3) publish_date 若提供了 YYYY-MM-DD 则沿用，否则填 N/A\n4) 只输出严格 JSON 数组（不要解释、不要 markdown、不要代码块）\n\n上下文：\n- region_name: {loc_name}\n- chain_name: {chain_name}\n- node_name: {node_name}\n- chain_node: {chain_node}\n- chain_ids: {chain_ids}\n- node_ids: {node_ids}\n- keyword_group: {group_name}\n- keyword: {keyword}\n\n原始新闻：\n- title: {title}\n- summary: {summary}\n- entities: {entities}\n- key_event: {key_event}\n- platform: {platform}\n- link: {link}\n- publish_date: {publish_date}\n\n请输出如下 JSON 数组结构（Key 必须完全一致，不允许缺失；输出 1-2 条；不确定输出 []）：\n[\n  {{\n    \"clue_name\": \"项目/线索名称（如扩产/融资/新建基地）\",\n    \"company_name\": \"企业全称\",\n    \"investment_amount\": \"投资金额/融资金额/N/A\",\n    \"land_requirement\": \"用地/厂房需求/N/A\",\n    \"investment_layout\": \"异地布局/扩张方向/N/A\",\n    \"content\": \"200字以内项目详情（尽量引用摘要的事实）/N/A\",\n    \"product_details\": \"主营产品/技术/N/A\",\n    \"company_intro\": \"企业简介/N/A\",\n    \"company_customers\": \"主要客户/N/A\",\n    \"company_revenue\": \"营收规模/N/A\",\n    \"shareholder_background\": \"母公司/股东背景/N/A\",\n    \"investment_plan_docs\": \"[]\",\n    \"public_sentiment_link\": \"{link}\",\n    \"publish_date\": \"{publish_date}\",\n    \"registration_place\": \"注册地/总部所在地/N/A\",\n    \"recommend_reason\": \"结合证据给出跟进理由（含关键信息点）/N/A\",\n    \"chain_name\": \"{chain_name}\",\n    \"region_name\": \"{loc_name}\",\n    \"node_name\": \"{node_name}\",\n    \"chain_node\": \"{chain_node}\",\n    \"chain_ids\": \"{chain_ids}\",\n    \"node_ids\": \"{node_ids}\"\n  }}\n]"""

    def log(self, message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] {message}"
        print(log_entry)
        self.logs.append(log_entry)
        if len(self.logs) > 200:
            self.logs.pop(0)

    def get_logs(self):
        return self.logs
        
    def get_leads(self):
        return self.leads

    def get_raw_news(self):
        return self.raw_news

    def _try_load_json(self, path):
        p = (path or "").strip()
        if not p:
            return None
        if not os.path.isabs(p):
            p = os.path.join(self.base_dir, p)
        if not os.path.exists(p):
            return None
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
        except Exception:
            return None
        return None

    def _load_run_config(self):
        candidates = [
            os.path.join(self.base_dir, "run_config.json"),
            os.path.join(os.path.dirname(self.base_dir), "run_config.json"),
            os.path.join(os.getcwd(), "run_config.json"),
        ]
        for p in candidates:
            cfg = self._try_load_json(p)
            if cfg:
                return cfg
        return {}

    def _cfg_raw(self, env_key, json_key, default):
        v = os.getenv(env_key)
        if v is not None and str(v).strip() != "":
            return str(v).strip()
        cfg = self.run_config or {}
        if isinstance(cfg, dict) and json_key in cfg:
            return cfg.get(json_key)
        return default

    def parse_excel_chains(self, file_path):
        """解析上传的 Excel 文件中的 Sheet 名称作为产业链"""
        try:
            self.nodes_file = str(file_path or "").strip()
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            chains = wb.sheetnames
            self.config["industry_chains"] = chains
            wb.close()
            self._build_node_meta_from_nodes_file()
            self.log(f"成功解析Excel文件，获取到产业链: {', '.join(chains)}")
            return chains
        except Exception as e:
            self.log(f"解析Excel文件失败: {str(e)}")
            return []

    def update_config(self, new_config):
        self.config.update(new_config)
        self.log(f"配置已更新: 目标落商地={self.config['location_name']}(ID:{self.config['location_id']}), 节点数={len(self.config['nodes'])}")

    def stop_task(self):
        self.log("接收到停止信号，正在停止任务...")
        self._stop_event.set()

    def _http_get_text(self, url, timeout=12):
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        }
        try:
            r = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
            if r.status_code >= 400:
                return ""
            r.encoding = r.apparent_encoding or r.encoding
            return r.text or ""
        except Exception:
            return ""

    def _extract_urls_from_html(self, html, limit=5):
        if not html:
            return []
        return self._extract_urls_from_html_engine(html, limit=limit, engine="generic")

    def _is_static_url(self, url):
        u = (url or "").strip()
        if not u.startswith("http"):
            return True
        try:
            p = urlparse(u)
            path = (p.path or "").lower()
            host = (p.netloc or "").lower()
        except Exception:
            return True
        if "bcebos.com" in host and "/static/" in path:
            return True
        static_ext = (
            ".css",
            ".js",
            ".png",
            ".jpg",
            ".jpeg",
            ".gif",
            ".svg",
            ".webp",
            ".ico",
            ".woff",
            ".woff2",
            ".ttf",
            ".eot",
            ".map",
            ".mp4",
            ".mp3",
            ".m3u8",
        )
        return any(path.endswith(ext) for ext in static_ext)

    def _is_blacklisted_url(self, url):
        u = (url or "").strip()
        if not u.startswith("http"):
            return True
        try:
            host = (urlparse(u).netloc or "").lower()
        except Exception:
            return False
        if not host:
            return True
        blocked_hosts = (
            "mlb.com",
            "nypost.com",
            "nytimes.com",
            "theathletic.com",
            "heavy.com",
            "vercel.app",
            "channellineups.com",
            "channellists.com",
        )
        return any(host == h or host.endswith("." + h) for h in blocked_hosts)

    def _has_chinese(self, text):
        s = (text or "").strip()
        if not s:
            return False
        return bool(re.search(r"[\u4e00-\u9fff]", s))

    def _looks_relevant(self, text, chain_name, node_name, keyword):
        s = (text or "")
        if not self._has_chinese(s):
            return False
        chain = str(chain_name or "").strip()
        node = str(node_name or "").strip()
        kw = str(keyword or "").strip()
        hit = 0
        if chain and chain in s:
            hit += 1
        if node and node in s:
            hit += 1
        if kw and kw in s:
            hit += 1
        return hit >= 1

    def _normalize_result_url(self, url):
        u = unescape((url or "").strip())
        if not u.startswith("http"):
            return ""
        try:
            p = urlparse(u)
            host = (p.netloc or "").lower()
        except Exception:
            return u
        if "bing.com" in host and (p.path or "").startswith("/ck/"):
            qs = parse_qs(p.query or "")
            raw = (qs.get("u") or [""])[0]
            raw = unescape(str(raw or "")).strip()
            if raw.startswith("a1"):
                b64 = raw[2:]
                pad = "=" * ((4 - (len(b64) % 4)) % 4)
                try:
                    decoded = base64.b64decode((b64 + pad).encode("utf-8"), validate=False).decode("utf-8", errors="ignore")
                    decoded = decoded.strip()
                    if decoded.startswith("http"):
                        return decoded
                except Exception:
                    pass
            return u
        if "baidu.com" in host and (p.path or "").startswith("/link"):
            try:
                headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
                    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                }
                r = requests.get(u, headers=headers, timeout=10, allow_redirects=True)
                if r.status_code >= 400:
                    return ""
                return (r.url or "").strip()
            except Exception:
                return ""
        return u

    def _extract_urls_from_html_engine(self, html, limit=5, engine="generic"):
        if not html:
            return []
        if engine == "bing":
            candidates = re.findall(r'(?is)<li[^>]+class="b_algo"[^>]*>.*?<h2[^>]*>.*?<a[^>]+href="([^"]+)"', html)
        elif engine == "baidu":
            candidates = re.findall(r'(?is)<h3[^>]*class="t[^"]*"[^>]*>.*?<a[^>]+href="([^"]+)"', html)
        else:
            candidates = re.findall(r'href="(https?://[^"]+)"', html)

        links = []
        for m in candidates:
            u = unescape((m or "").strip())
            if not u.startswith("http"):
                continue
            u = self._normalize_result_url(u) or ""
            if not u:
                continue
            if self._is_static_url(u):
                continue
            if self._is_blacklisted_url(u):
                continue
            if u not in links:
                links.append(u)
            if len(links) >= limit:
                break
        return links[:limit]

    def _search_bing(self, query, limit=5):
        url = "https://www.bing.com/search?q=" + quote_plus(query) + "&setlang=zh-cn&cc=CN"
        html = self._http_get_text(url, timeout=12)
        return self._extract_urls_from_html_engine(html, limit=limit, engine="bing")

    def _search_baidu(self, query, limit=5):
        url = "https://www.baidu.com/s?wd=" + quote_plus(query)
        html = self._http_get_text(url, timeout=12)
        return self._extract_urls_from_html_engine(html, limit=limit, engine="baidu")

    def _html_to_text(self, html):
        if not html:
            return ""
        src = re.sub(r"(?is)<(script|style|noscript).*?>.*?</\1>", " ", html)
        src = re.sub(r"(?is)<br\\s*/?>", "\n", src)
        src = re.sub(r"(?is)</p\\s*>", "\n", src)
        src = re.sub(r"(?is)<[^>]+>", " ", src)
        src = unescape(src)
        src = re.sub(r"[ \t\r\f\v]+", " ", src)
        src = re.sub(r"\n{2,}", "\n", src)
        return src.strip()

    def _deepseek_chat(self, prompt):
        if not self.deepseek_api_key:
            raise RuntimeError("未配置 deepseek_api_key，无法使用 AI 联网分析")
        base = (self.deepseek_api_base_url or "https://api.deepseek.com").rstrip("/")
        url = base + "/v1/chat/completions"
        headers = {
            "Authorization": f"Bearer {self.deepseek_api_key}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": self.deepseek_model or "deepseek-chat",
            "messages": [
                {"role": "system", "content": "你是一个严格的数据抽取助手。只输出可解析的 JSON 数组，不要解释，不要 markdown，不要代码块标记。"},
                {"role": "user", "content": str(prompt)},
            ],
            "temperature": 0.2,
        }
        r = requests.post(url, headers=headers, json=payload, timeout=self.deepseek_timeout_seconds)
        r.raise_for_status()
        data = r.json() or {}
        choices = data.get("choices") or []
        msg = (choices[0] or {}).get("message") if choices else {}
        content = (msg or {}).get("content") or ""
        return str(content)

    def _parse_json_array(self, text):
        if not text:
            return []
        s = text.strip()
        if "```" in s:
            m = re.search(r"```(?:json)?\\s*([\\s\\S]*?)\\s*```", s, re.I)
            if m:
                s = (m.group(1) or "").strip()
        start, end = s.find("["), s.rfind("]")
        if start != -1 and end != -1 and end >= start:
            s = s[start : end + 1]
        s = re.sub(r",\\s*([\\]}])", r"\\1", s)
        s = s.replace("“", "\"").replace("”", "\"")
        try:
            out = json.loads(s)
            return out if isinstance(out, list) else []
        except Exception:
            return []

    def _invalid_company_name(self, name):
        t = (name or "").strip()
        if not t:
            return True
        bad_tokens = ["某公司", "某企业", "签约企业", "未披露", "相关企业", "多家企业", "某真实"]
        for b in bad_tokens:
            if b in t:
                return True
        if "公司" not in t:
            return True
        suffix_ok = any(t.endswith(x) for x in ["有限公司", "股份有限公司", "有限责任公司", "集团有限公司"])
        return not suffix_ok

    def _check_url_ok(self, url):
        u = (url or "").strip()
        if not u.startswith("http"):
            return False
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            }
            r = requests.get(u, headers=headers, timeout=10, allow_redirects=True)
            return r.status_code < 400
        except Exception:
            return False

    def _build_node_meta_from_nodes_file(self):
        self.node_meta = {}
        p = (self.nodes_file or "").strip()
        if not p or not os.path.exists(p):
            return
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception:
            return
        try:
            for sheet in (wb.worksheets or []):
                chain_name = str(sheet.title or "").strip()
                if not chain_name:
                    continue
                header = [sheet.cell(row=1, column=c).value for c in range(1, (sheet.max_column or 0) + 1)]
                header_norm = [str(v).strip() if v is not None else "" for v in header]

                def find_col(name, fallback_idx):
                    for i, hv in enumerate(header_norm, start=1):
                        if hv == name:
                            return i
                    return fallback_idx

                name_col = find_col("名称", 2)
                chain_ids_col = find_col("chain_ids", 0)
                node_ids_col = find_col("node_ids", 0)
                for r in range(2, (sheet.max_row or 0) + 1):
                    v = sheet.cell(row=r, column=name_col).value
                    node_name = str(v).strip() if v is not None else ""
                    if not node_name or node_name == "名称" or node_name == chain_name:
                        continue
                    chain_ids_v = sheet.cell(row=r, column=chain_ids_col).value if chain_ids_col else None
                    node_ids_v = sheet.cell(row=r, column=node_ids_col).value if node_ids_col else None
                    self.node_meta[(chain_name, node_name)] = {
                        "chain_ids": str(chain_ids_v).strip() if chain_ids_v is not None else "",
                        "node_ids": str(node_ids_v).strip() if node_ids_v is not None else "",
                    }
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def get_nodes_for_chain(self, chain_name):
        chain = str(chain_name or "").strip()
        if not chain:
            return []
        p = (self.nodes_file or "").strip()
        if not p or not os.path.exists(p):
            return [chain]
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            if chain not in wb.sheetnames:
                try:
                    wb.close()
                except Exception:
                    pass
                return [chain]
            ws = wb[chain]
            header = [ws.cell(row=1, column=c).value for c in range(1, (ws.max_column or 0) + 1)]
            header_norm = [str(v).strip() if v is not None else "" for v in header]
            name_col = 2
            for i, hv in enumerate(header_norm, start=1):
                if hv == "名称":
                    name_col = i
                    break
            nodes = []
            for r in range(2, (ws.max_row or 0) + 1):
                v = ws.cell(row=r, column=name_col).value
                node = str(v).strip() if v is not None else ""
                if not node or node == "名称" or node == chain:
                    continue
                if node not in nodes:
                    nodes.append(node)
            try:
                wb.close()
            except Exception:
                pass
            if not nodes:
                return [chain]
            limit = int(self.config.get("max_nodes_per_chain") or 0)
            if limit and limit > 0:
                nodes = nodes[:limit]
            return nodes
        except Exception:
            return [chain]

    def _build_lead_prompt(self, chain_name, node_name, group_name, keyword, url, article_text, meta):
        loc_name = self.config.get("location_name") or "未披露"
        today = datetime.now().strftime("%Y-%m-%d")
        t = (article_text or "").strip()
        if len(t) > 6000:
            t = t[:6000]
        chain_ids = str((meta or {}).get("chain_ids") or "")
        node_ids = str((meta or {}).get("node_ids") or "")
        chain_node = f"{chain_name}-{node_name}"
        return f"""你现在是一名“招商线索穿透分析师”。你将获得一个网页正文节选，请基于我提供的文字信息输出招商线索。\n\n硬性规则：\n1) company_name 必须是真实工商企业全称，且必须以“有限公司/股份有限公司/有限责任公司/集团有限公司”结尾；严禁使用“某公司/某企业/签约企业/未披露”等代称；无法确认输出 []\n2) public_sentiment_link 必须使用我提供的 url，且不得伪造\n3) investment_plan_docs 必须是 JSON 数组字符串（如：[] 或 [{{\"url\":\"...\",\"name\":\"...\"}}]），没有则输出 []\n4) 所有字段必须输出；未知填写 \"N/A\"\n5) 只输出严格 JSON 数组（不要解释、不要 markdown、不要代码块）\n\n上下文：\n- region_name: {loc_name}\n- chain_name: {chain_name}\n- node_name: {node_name}\n- chain_node: {chain_node}\n- chain_ids: {chain_ids}\n- node_ids: {node_ids}\n- keyword_group: {group_name}\n- keyword: {keyword}\n- url: {url}\n- crawl_date: {today}\n\n网页正文（节选）：\n{t}\n\n请输出如下 JSON 数组结构（Key 必须完全一致，不允许缺失；输出 1-3 条；不确定输出 []）：\n[\n  {{\n    \"clue_name\": \"项目/线索名称（如扩产/融资/新建基地）\",\n    \"company_name\": \"企业全称\",\n    \"investment_amount\": \"投资金额/融资金额/N/A\",\n    \"land_requirement\": \"用地/厂房需求/N/A\",\n    \"investment_layout\": \"异地布局/扩张方向/N/A\",\n    \"content\": \"200字以内项目详情/N/A\",\n    \"product_details\": \"主营产品/技术/N/A\",\n    \"company_intro\": \"企业简介/N/A\",\n    \"company_customers\": \"主要客户/N/A\",\n    \"company_revenue\": \"营收规模/N/A\",\n    \"shareholder_background\": \"母公司/股东背景/N/A\",\n    \"investment_plan_docs\": \"[]\",\n    \"public_sentiment_link\": \"{url}\",\n    \"publish_date\": \"YYYY-MM-DD或N/A\",\n    \"registration_place\": \"注册地/总部所在地/N/A\",\n    \"recommend_reason\": \"证据来源与跟进理由/N/A\",\n    \"chain_name\": \"{chain_name}\",\n    \"region_name\": \"{loc_name}\",\n    \"node_name\": \"{node_name}\",\n    \"chain_node\": \"{chain_node}\",\n    \"chain_ids\": \"{chain_ids}\",\n    \"node_ids\": \"{node_ids}\"\n  }}\n]"""

    def _iter_keywords(self, keywords):
        ks = keywords or []
        n = max(1, min(6, len(ks)))
        return ks[:n]

    def run_task(self):
        """核心爬取和分析任务逻辑"""
        self.is_running = True
        self._stop_event.clear()
        self.log("🚀 全自动招商线索挖掘任务启动（先采集新闻 -> 再穿透成线索）")
        
        chains = self.config.get("industry_chains", [])
        if not chains:
            self.log("⚠️ 警告：未检测到产业链配置，将使用默认产业链：新能源、新材料")
            chains = ["新能源", "新材料"]
            
        nodes = self.config.get("nodes", ["主节点"])
        loc_name = self.config.get("location_name", "未知")
        loc_id = self.config.get("location_id", "0")

        self.log(f"⚙️ 运行参数 -> 目标落户地: {loc_name}({loc_id}), 分布式节点: {', '.join(nodes)}")
        if not self.deepseek_api_key:
            self.log("❌ 未检测到 DeepSeek API Key：请在 run_config.json 配置 deepseek_api_key，或设置环境变量 DEEPSEEK_API_KEY")
            self.is_running = False
            return

        self.leads = []
        self.raw_news = []
        self.stats = {"total_crawled": 0, "valid_leads": 0, "raw_news": 0}

        news_source = str(self.config.get("news_source") or "raw_news_csv").strip()
        raw_news_rows = []
        raw_news_csv = ""
        if news_source == "raw_news_csv":
            raw_news_csv = self._find_raw_news_csv()
            if not raw_news_csv:
                self.log("❌ 未找到 raw_news_leads.csv。请先运行上级目录的 crawl_raw_news.py 采集原始新闻，再回到本系统进行穿透分析。")
                self.is_running = False
                return
            raw_news_rows = self._load_raw_news_rows_from_csv(raw_news_csv)
            self.log(f"🗞️ 已加载原始新闻 {len(raw_news_rows)} 条（来源：{raw_news_csv}）")
        
        for chain in chains:
            if self._stop_event.is_set():
                break
                
            nodes_for_chain = self.get_nodes_for_chain(chain)
            if not nodes_for_chain:
                nodes_for_chain = [chain]
            self.log(f"\n========== 开始处理产业链: {chain}（节点数 {len(nodes_for_chain)}） ==========")

            for node_name in nodes_for_chain:
                if self._stop_event.is_set():
                    break
                meta = self.node_meta.get((chain, node_name), {}) if self.node_meta else {}
                max_raw = int(self.config.get("max_raw_news_per_node") or 30)
                chain_node_label = f"{chain}-{node_name}" if node_name and node_name != chain else str(chain)

                if news_source == "raw_news_csv":
                    node_news = []
                    for r in raw_news_rows:
                        label = self._normalize_text(r.get("所属产业") or r.get("所属产业链") or "")
                        if not label:
                            continue
                        if chain_node_label in label or label.startswith(str(chain)):
                            node_news.append(r)
                    if not node_news:
                        continue
                    self.log(f"🧾 [{chain}-{node_name}] 使用原始新闻穿透：命中 {len(node_news)} 条（取前 {max_raw} 条）")

                    for idx, raw_row in enumerate(node_news[:max_raw], start=1):
                        if self._stop_event.is_set():
                            break
                        link = self._normalize_text((raw_row or {}).get("新闻链接") or "")
                        title = self._normalize_text((raw_row or {}).get("新闻标题") or "")
                        if not link:
                            continue
                        self.raw_news.append(
                            {
                                "chain_name": chain,
                                "node_name": node_name,
                                "public_sentiment_link": link,
                                "title": title,
                                "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            }
                        )
                        self.stats["raw_news"] += 1

                        group_name = self._normalize_text((raw_row or {}).get("搜索组别") or "原始新闻")
                        keyword = self._normalize_text((raw_row or {}).get("关键事件") or "")
                        self.log(f"📄 [{idx}/{min(len(node_news), max_raw)}] 穿透分析：{title}")

                        prompt = self._build_penetration_prompt_from_raw_news(chain, node_name, group_name, keyword, raw_row, meta)
                        try:
                            resp = self._deepseek_chat(prompt)
                        except Exception as e:
                            self.log(f"❌ AI 调用失败：{str(e)}")
                            continue
                        data = self._parse_json_array(resp)
                        self.stats["total_crawled"] += 1
                        if not data:
                            continue
                        added = 0
                        for item in data:
                            if not isinstance(item, dict):
                                continue
                            company = (item.get("company_name") or "").strip()
                            if self._invalid_company_name(company):
                                continue
                            if not self._check_url_ok(link):
                                continue

                            def _ns(v):
                                if v is None:
                                    return "N/A"
                                t = str(v).strip()
                                return t if t else "N/A"

                            investment_plan_docs = item.get("investment_plan_docs")
                            if isinstance(investment_plan_docs, list):
                                try:
                                    investment_plan_docs = json.dumps(investment_plan_docs[:1], ensure_ascii=False)
                                except Exception:
                                    investment_plan_docs = "[]"
                            elif isinstance(investment_plan_docs, dict):
                                try:
                                    investment_plan_docs = json.dumps([investment_plan_docs], ensure_ascii=False)
                                except Exception:
                                    investment_plan_docs = "[]"
                            else:
                                investment_plan_docs = str(investment_plan_docs or "").strip() or "[]"

                            row = {
                                "clue_name": _ns(item.get("clue_name")),
                                "company_name": company,
                                "investment_amount": _ns(item.get("investment_amount")),
                                "land_requirement": _ns(item.get("land_requirement")),
                                "investment_layout": _ns(item.get("investment_layout")),
                                "content": _ns(item.get("content")),
                                "product_details": _ns(item.get("product_details")),
                                "company_intro": _ns(item.get("company_intro")),
                                "company_customers": _ns(item.get("company_customers")),
                                "company_revenue": _ns(item.get("company_revenue")),
                                "shareholder_background": _ns(item.get("shareholder_background")),
                                "investment_plan_docs": investment_plan_docs,
                                "public_sentiment_link": link,
                                "publish_date": _ns(item.get("publish_date")),
                                "registration_place": _ns(item.get("registration_place")),
                                "recommend_reason": _ns(item.get("recommend_reason")),
                                "chain_name": str(chain),
                                "region_name": str(loc_name),
                                "node_name": str(node_name),
                                "chain_node": f"{chain}-{node_name}",
                                "chain_ids": str(meta.get("chain_ids") or ""),
                                "node_ids": str(meta.get("node_ids") or ""),
                            }
                            dedup_key = (row["company_name"], row["clue_name"], row["chain_node"], row["public_sentiment_link"])
                            if any(
                                (x.get("company_name"), x.get("clue_name"), x.get("chain_node"), x.get("public_sentiment_link")) == dedup_key
                                for x in self.leads
                            ):
                                continue
                            self.leads.insert(0, row)
                            added += 1
                            self.stats["valid_leads"] += 1
                        if added:
                            self.log(f"🌟 新增有效线索 {added} 条")
                else:
                    self.log(f"🧾 [{chain}-{node_name}] 使用搜索引擎抓取穿透（备用模式）")
                    per_query = int(self.config.get("search_results_per_query") or 5)
                    max_urls = int(self.config.get("max_urls_per_chain") or 20)
                    seen_urls = set()

                    for group_name, keywords in self.keyword_groups.items():
                        if self._stop_event.is_set() or len(seen_urls) >= max_urls:
                            break
                        for kw in self._iter_keywords(keywords):
                            if self._stop_event.is_set() or len(seen_urls) >= max_urls:
                                break
                            q = f"{chain} {node_name} {kw} 企业 项目 新闻"
                            self.log(f"🔎 联网搜索：{q}")
                            found = self._search_bing(q, limit=per_query)
                            if not found:
                                found = self._search_baidu(q, limit=per_query)
                            if not found:
                                continue

                            for u in found:
                                if self._stop_event.is_set() or len(seen_urls) >= max_urls:
                                    break
                                if not u or u in seen_urls:
                                    continue
                                if self._is_static_url(u) or self._is_blacklisted_url(u):
                                    continue
                                seen_urls.add(u)
                                if not self._check_url_ok(u):
                                    continue
                                html = self._http_get_text(u, timeout=12)
                                article_text = self._html_to_text(html)
                                if not article_text or len(article_text) < 200:
                                    continue
                                if not self._looks_relevant(article_text, chain, node_name, kw):
                                    continue

                                self.raw_news.append(
                                    {"chain_name": chain, "node_name": node_name, "public_sentiment_link": u, "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
                                )
                                self.stats["raw_news"] += 1
                                self.log(f"📄 [{len(seen_urls)}/{max_urls}] 抽取线索：{u}")

                                prompt = self._build_lead_prompt(chain, node_name, group_name, kw, u, article_text, meta)
                                try:
                                    resp = self._deepseek_chat(prompt)
                                except Exception as e:
                                    self.log(f"❌ AI 调用失败：{str(e)}")
                                    continue
                                data = self._parse_json_array(resp)
                                self.stats["total_crawled"] += 1
                                if not data:
                                    continue
                                added = 0
                                for item in data:
                                    if not isinstance(item, dict):
                                        continue
                                    company = (item.get("company_name") or "").strip()
                                    if self._invalid_company_name(company):
                                        continue
                                    if not self._check_url_ok(u):
                                        continue

                                    def _ns(v):
                                        if v is None:
                                            return "N/A"
                                        t = str(v).strip()
                                        return t if t else "N/A"

                                    investment_plan_docs = item.get("investment_plan_docs")
                                    if isinstance(investment_plan_docs, list):
                                        try:
                                            investment_plan_docs = json.dumps(investment_plan_docs[:1], ensure_ascii=False)
                                        except Exception:
                                            investment_plan_docs = "[]"
                                    elif isinstance(investment_plan_docs, dict):
                                        try:
                                            investment_plan_docs = json.dumps([investment_plan_docs], ensure_ascii=False)
                                        except Exception:
                                            investment_plan_docs = "[]"
                                    else:
                                        investment_plan_docs = str(investment_plan_docs or "").strip() or "[]"

                                    row = {
                                        "clue_name": _ns(item.get("clue_name")),
                                        "company_name": company,
                                        "investment_amount": _ns(item.get("investment_amount")),
                                        "land_requirement": _ns(item.get("land_requirement")),
                                        "investment_layout": _ns(item.get("investment_layout")),
                                        "content": _ns(item.get("content")),
                                        "product_details": _ns(item.get("product_details")),
                                        "company_intro": _ns(item.get("company_intro")),
                                        "company_customers": _ns(item.get("company_customers")),
                                        "company_revenue": _ns(item.get("company_revenue")),
                                        "shareholder_background": _ns(item.get("shareholder_background")),
                                        "investment_plan_docs": investment_plan_docs,
                                        "public_sentiment_link": u,
                                        "publish_date": _ns(item.get("publish_date")),
                                        "registration_place": _ns(item.get("registration_place")),
                                        "recommend_reason": _ns(item.get("recommend_reason")),
                                        "chain_name": str(chain),
                                        "region_name": str(loc_name),
                                        "node_name": str(node_name),
                                        "chain_node": f"{chain}-{node_name}",
                                        "chain_ids": str(meta.get("chain_ids") or ""),
                                        "node_ids": str(meta.get("node_ids") or ""),
                                    }
                                    dedup_key = (row["company_name"], row["clue_name"], row["chain_node"], row["public_sentiment_link"])
                                    if any(
                                        (x.get("company_name"), x.get("clue_name"), x.get("chain_node"), x.get("public_sentiment_link")) == dedup_key
                                        for x in self.leads
                                    ):
                                        continue
                                    self.leads.insert(0, row)
                                    added += 1
                                    self.stats["valid_leads"] += 1
                                if added:
                                    self.log(f"🌟 新增有效线索 {added} 条")
            
        if self._stop_event.is_set():
            self.log("⏹️ 任务已手动终止")
        else:
            self.log("✅ 所有产业链挖掘任务执行完毕！")
            
        self.is_running = False
